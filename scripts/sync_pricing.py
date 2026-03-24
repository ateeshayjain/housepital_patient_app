#!/usr/bin/env python3
"""Sync Housepital Unified Pricing Master Excel into equipment_catalog.json."""

import json
import openpyxl
import sys
from pathlib import Path

EXCEL_PATH = Path.home() / "Downloads" / "Housepital_Unified_Pricing_Master.xlsx"
JSON_PATH = Path.home() / "housepital_patient_app" / "assets" / "equipment_catalog.json"

def main():
    # Load Excel
    wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)
    ws = wb["Products"]

    # Build lookup: ID -> row data
    excel_data = {}
    for row in range(2, ws.max_row + 1):
        item_id = ws.cell(row=row, column=1).value
        if not item_id:
            continue
        item_id = str(item_id).strip()

        sale_price = ws.cell(row=row, column=7).value  # G: Sale Price
        mrp = ws.cell(row=row, column=8).value  # H: MRP
        avail_rent_raw = ws.cell(row=row, column=9).value  # I: Available for Rent
        rental_7d = ws.cell(row=row, column=10).value  # J: Rental 7 Days
        rental_30d = ws.cell(row=row, column=11).value  # K: Rental 30 Days
        use_case = ws.cell(row=row, column=22).value  # V: Use Case / Condition

        # Parse available_for_rent
        avail_rent = False
        if avail_rent_raw is not None:
            if isinstance(avail_rent_raw, bool):
                avail_rent = avail_rent_raw
            elif isinstance(avail_rent_raw, str):
                avail_rent = avail_rent_raw.strip().lower() in ("yes", "true", "1")

        excel_data[item_id] = {
            "sale_price": float(sale_price) if sale_price is not None else None,
            "mrp": float(mrp) if mrp is not None else None,
            "available_for_rent": avail_rent,
            "rental_7d": float(rental_7d) if rental_7d is not None else None,
            "rental_30d": float(rental_30d) if rental_30d is not None else None,
            "use_case": str(use_case).strip() if use_case else None,
        }

    print(f"Loaded {len(excel_data)} items from Excel")

    # Load JSON
    with open(JSON_PATH, "r", encoding="utf-8") as f:
        catalog = json.load(f)

    print(f"Loaded {len(catalog)} items from JSON")

    updated = 0
    for item in catalog:
        item_id = item.get("id")
        if item_id not in excel_data:
            continue

        xd = excel_data[item_id]
        changed = False

        # MRP (new field)
        old_mrp = item.get("mrp")
        new_mrp = xd["mrp"]
        if old_mrp != new_mrp:
            item["mrp"] = new_mrp
            changed = True

        # Sale Price -> price (overwrite if Excel has a value)
        if xd["sale_price"] is not None:
            old_price = item.get("price")
            if old_price != xd["sale_price"]:
                item["price"] = xd["sale_price"]
                changed = True

        # Available for Rent
        old_rent = item.get("available_for_rent")
        if old_rent != xd["available_for_rent"]:
            item["available_for_rent"] = xd["available_for_rent"]
            changed = True

        # Rental tiers
        if xd["rental_7d"] is not None or xd["rental_30d"] is not None:
            tiers = item.get("rental_tiers", {}) or {}
            if xd["rental_7d"] is not None:
                tiers["7d"] = xd["rental_7d"]
            if xd["rental_30d"] is not None:
                tiers["30d"] = xd["rental_30d"]
            if tiers != item.get("rental_tiers"):
                item["rental_tiers"] = tiers
                changed = True

        # Use Case (new field)
        old_uc = item.get("use_case")
        new_uc = xd["use_case"]
        if old_uc != new_uc:
            item["use_case"] = new_uc
            changed = True

        if changed:
            updated += 1

    # Write JSON
    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(catalog, f, indent=2, ensure_ascii=False)
        f.write("\n")

    print(f"Updated {updated} items in JSON")
    print("Done!")

if __name__ == "__main__":
    main()
