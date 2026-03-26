#!/usr/bin/env python3
"""
Sync Excel data from Housepital_Unified_Pricing_Master_Fixed.xlsx to equipment_catalog.json.
Includes the new SPOC field and fuzzy image matching.
"""

import json
import os
import re
import openpyxl

EXCEL_PATH = '/Users/ateeshayjain/Desktop/Housepital/App Data - Scope of Service/Housepital_Unified_Pricing_Master_Fixed.xlsx'
IMAGES_DIR = '/Users/ateeshayjain/housepital_patient_app/assets/images/products'
OUTPUT_PATH = '/Users/ateeshayjain/housepital_patient_app/assets/equipment_catalog.json'


def normalize_image_filename(filename):
    """Remove numeric prefix and extension, replace _ with space."""
    name = os.path.splitext(filename)[0]
    # Remove leading number prefix like 0001_
    name = re.sub(r'^\d+_', '', name)
    # Replace underscores with spaces
    name = name.replace('_', ' ')
    return name.strip().lower()


def normalize_for_match(s):
    """Normalize a string for fuzzy matching: lowercase, remove special chars, collapse spaces."""
    s = s.lower()
    s = re.sub(r'[^a-z0-9 ]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def build_image_map(images_dir):
    """Build a map of normalized name -> first matching filename."""
    image_map = {}
    files = sorted(os.listdir(images_dir))
    for f in files:
        if not f.lower().endswith(('.jpg', '.png', '.jpeg')):
            continue
        normalized = normalize_image_filename(f)
        norm_key = normalize_for_match(normalized)
        if norm_key not in image_map:
            image_map[norm_key] = f
    return image_map


def find_image_for_item(item_name, image_map):
    """Find the best matching image for an item name."""
    name_norm = normalize_for_match(item_name)
    # Exact normalized match
    if name_norm in image_map:
        return image_map[name_norm]
    # Try substring matching - if image name contains the item name or vice versa
    for key, filename in image_map.items():
        if key == name_norm or key.startswith(name_norm) or name_norm.startswith(key):
            return filename
    return None


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['Products']

    # Read headers from first row
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(cell.value)

    print(f"Headers found: {headers}")

    # Build image map
    image_map = build_image_map(IMAGES_DIR)
    print(f"Image map has {len(image_map)} unique entries")

    # Map Excel header names to JSON field names (matching existing app schema)
    header_to_json = {
        'ID': 'id',
        'Item Name': 'name',
        'SPOC': 'spoc',
        'Brand': 'brand',
        'Category': 'category',
        'Status': 'status',
        'Available for Sale': 'available_for_sale',
        'Sale Price': 'price',
        'MRP': 'mrp',
        'Available for Rent': 'available_for_rent',
        'Rental 30 Days': 'rental_price',
        'Tariff Name': 'tariff_name',
        'Parent Product ID': 'parent_product_id',
        'Variant Type': 'variant_type',
        'Variant Value': 'variant_value',
        'Description': 'description',
        'How to Use': 'how_to_use',
        'Key Features': 'key_features',
        'Ideal For': 'ideal_for',
        'YouTube Tutorial Link': 'youtube_url',
        'FAQs': 'faqs',
        'Use Case / Condition': 'use_case',
    }

    # Build header index mapping
    col_map = {}
    for i, h in enumerate(headers):
        if h and h.strip() in header_to_json:
            col_map[i] = header_to_json[h.strip()]

    items = []
    matched_images = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # Skip empty rows
            continue

        item = {}
        for i, value in enumerate(row):
            if i in col_map:
                item[col_map[i]] = value

        # Ensure required fields
        if not item.get('name'):
            continue

        # Convert boolean fields
        for bool_field in ['available_for_sale', 'available_for_rent']:
            val = item.get(bool_field)
            if isinstance(val, str):
                item[bool_field] = val.strip().lower() in ('true', 'yes', '1')
            elif isinstance(val, (int, float)):
                item[bool_field] = bool(val)
            elif val is None:
                item[bool_field] = False

        # Convert numeric fields
        for num_field in ['price', 'mrp', 'rental_price']:
            if num_field in item and item[num_field] is not None:
                try:
                    item[num_field] = float(item[num_field])
                except (ValueError, TypeError):
                    item[num_field] = None

        # Set rental_tiers as empty dict (not in Excel)
        item['rental_tiers'] = {}

        # Match image
        item_name = item.get('name', '')
        matched_file = find_image_for_item(item_name, image_map)
        if matched_file:
            item['image_url'] = f'assets/images/products/{matched_file}'
            matched_images += 1
        else:
            item['image_url'] = None

        items.append(item)

    print(f"Total items: {len(items)}")
    print(f"Matched images: {matched_images}")

    # Show some unmatched items for debugging
    unmatched = [it['name'] for it in items if it['image_url'] is None][:10]
    if unmatched:
        print(f"Sample unmatched: {unmatched}")

    # Write JSON
    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(items, f, indent=2, ensure_ascii=False, default=str)

    print(f"Written to {OUTPUT_PATH}")


if __name__ == '__main__':
    main()
